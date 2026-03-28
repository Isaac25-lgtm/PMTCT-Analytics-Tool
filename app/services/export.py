"""
In-memory export helpers for PDF, XLSX, and CSV report downloads.

This module intentionally keeps optional dependencies lazy so the main app can
start even when export libraries are not installed yet.
"""

from __future__ import annotations

import csv
import io
from datetime import datetime, timezone
from typing import Any, Optional


class ExportDependencyError(RuntimeError):
    """Raised when an optional export dependency is unavailable."""


def _timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S")


def _safe_name(value: str) -> str:
    cleaned = "".join(character if character.isalnum() else "_" for character in value)
    return cleaned.strip("_")[:40] or "org_unit"


class PDFExporter:
    """Generate PDF reports in memory."""

    def _load(self) -> dict[str, Any]:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        except ImportError as exc:
            raise ExportDependencyError(
                "reportlab is required for PDF export. Install requirements-export.txt."
            ) from exc

        return {
            "colors": colors,
            "A4": A4,
            "ParagraphStyle": ParagraphStyle,
            "getSampleStyleSheet": getSampleStyleSheet,
            "mm": mm,
            "Paragraph": Paragraph,
            "SimpleDocTemplate": SimpleDocTemplate,
            "Spacer": Spacer,
            "Table": Table,
            "TableStyle": TableStyle,
        }

    def _styles(self, rl: dict[str, Any]) -> dict[str, Any]:
        styles = rl["getSampleStyleSheet"]()
        colors = rl["colors"]
        return {
            "title": rl["ParagraphStyle"](
                "ExportTitle",
                parent=styles["Title"],
                fontSize=18,
                textColor=colors.HexColor("#006B3F"),
                spaceAfter=8,
            ),
            "subtitle": rl["ParagraphStyle"](
                "ExportSubtitle",
                parent=styles["Normal"],
                fontSize=10,
                textColor=colors.HexColor("#64748B"),
                spaceAfter=14,
            ),
            "footer": rl["ParagraphStyle"](
                "ExportFooter",
                parent=styles["Normal"],
                fontSize=8,
                textColor=colors.HexColor("#64748B"),
                alignment=1,
            ),
        }

    def _table_style(self, rl: dict[str, Any]) -> Any:
        colors = rl["colors"]
        return rl["TableStyle"](
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#006B3F")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ]
        )

    def _build(self, title: str, subtitle: str, rows: list[list[Any]], widths: list[float]) -> bytes:
        rl = self._load()
        styles = self._styles(rl)
        buffer = io.BytesIO()
        document = rl["SimpleDocTemplate"](
            buffer,
            pagesize=rl["A4"],
            leftMargin=16 * rl["mm"],
            rightMargin=16 * rl["mm"],
            topMargin=18 * rl["mm"],
            bottomMargin=18 * rl["mm"],
        )
        table = rl["Table"](rows, colWidths=widths)
        table.setStyle(self._table_style(rl))
        story = [
            rl["Paragraph"]("PMTCT Triple Elimination", styles["title"]),
            rl["Paragraph"](title, styles["title"]),
            rl["Paragraph"](subtitle, styles["subtitle"]),
            table,
            rl["Spacer"](1, 16),
            rl["Paragraph"](
                f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
                styles["footer"],
            ),
        ]
        document.build(story)
        return buffer.getvalue()

    def generate_scorecard_pdf(
        self,
        indicators: list[dict[str, Any]],
        summary: dict[str, Any],
        org_unit: str,
        org_unit_name: Optional[str],
        period: str,
    ) -> bytes:
        rows = [
            ["Indicator", "ID", "Value", "Target", "Status"],
            ["Indicators reviewed", "", summary.get("total", 0), "", ""],
            ["Meeting target", "", summary.get("meeting_target", 0), "", ""],
            ["Score", "", f"{summary.get('score_pct', 0):.0f}%", "", ""],
        ]
        rows.extend(
            [
                indicator.get("name", ""),
                indicator.get("id", ""),
                indicator.get("formatted_value", "N/A"),
                f">= {indicator.get('target')}%" if indicator.get("target") is not None else "N/A",
                indicator.get("status", "unknown").title(),
            ]
            for indicator in indicators
        )
        return self._build(
            "WHO Validation Scorecard",
            f"{org_unit_name or org_unit} - {period}",
            rows,
            [72, 48, 64, 60, 60],
        )

    def generate_cascade_pdf(
        self,
        cascade_type: str,
        steps: list[dict[str, Any]],
        org_unit: str,
        org_unit_name: Optional[str],
        period: str,
    ) -> bytes:
        rows = [["Step", "Indicator", "Count", "Coverage"]]
        rows.extend(
            [index, step.get("name", ""), step.get("count", "N/A"), step.get("formatted_value", "N/A")]
            for index, step in enumerate(steps, start=1)
        )
        return self._build(
            f"{cascade_type.upper()} Cascade",
            f"{org_unit_name or org_unit} - {period}",
            rows,
            [42, 220, 70, 70],
        )

    def generate_supply_pdf(
        self,
        commodities: list[dict[str, Any]],
        org_unit: str,
        org_unit_name: Optional[str],
        period: str,
    ) -> bytes:
        rows = [["Commodity", "Stock on hand", "Consumed", "Stockout days", "Days of use", "Status"]]
        rows.extend(
            [
                commodity.get("commodity", ""),
                commodity.get("stock_on_hand", "N/A"),
                commodity.get("consumed", "N/A"),
                commodity.get("stockout_days", "N/A"),
                commodity.get("days_of_use", "N/A"),
                commodity.get("status", "unknown").title(),
            ]
            for commodity in commodities
        )
        return self._build(
            "Supply Chain Status",
            f"{org_unit_name or org_unit} - {period}",
            rows,
            [130, 65, 55, 65, 60, 50],
        )


class ExcelExporter:
    """Generate XLSX reports in memory."""

    def _load(self) -> dict[str, Any]:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Border, Font, PatternFill, Side
        except ImportError as exc:
            raise ExportDependencyError(
                "openpyxl is required for XLSX export. Install requirements-export.txt."
            ) from exc

        return {"Workbook": Workbook, "Border": Border, "Font": Font, "PatternFill": PatternFill, "Side": Side}

    def _write_sheet(self, workbook: Any, title: str, headers: list[str], rows: list[list[Any]], oxl: dict[str, Any]) -> None:
        sheet = workbook.active if workbook.active.title == "Sheet" else workbook.create_sheet(title)
        sheet.title = title
        sheet.append(headers)
        for row in rows:
            sheet.append(row)

        fill = oxl["PatternFill"](fill_type="solid", start_color="006B3F", end_color="006B3F")
        font = oxl["Font"](bold=True, color="FFFFFF")
        thin = oxl["Side"](style="thin", color="CBD5E1")
        border = oxl["Border"](left=thin, right=thin, top=thin, bottom=thin)

        for cell in sheet[1]:
            cell.fill = fill
            cell.font = font
            cell.border = border

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = border

        for column_cells in sheet.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max_len + 2, 40)

    def _apply_status_fills(self, sheet: Any, status_col: int, oxl: dict[str, Any]) -> None:
        """Apply traffic-light background fills to rows based on the status column."""
        fills = {
            "success": oxl["PatternFill"](fill_type="solid", start_color="D1FAE5", end_color="D1FAE5"),
            "warning": oxl["PatternFill"](fill_type="solid", start_color="FEF3C7", end_color="FEF3C7"),
            "danger": oxl["PatternFill"](fill_type="solid", start_color="FEE2E2", end_color="FEE2E2"),
        }
        for row in sheet.iter_rows(min_row=2):
            status_cell = row[status_col - 1] if len(row) >= status_col else None
            if not status_cell or not status_cell.value:
                continue
            status_val = str(status_cell.value).lower()
            fill = fills.get(status_val)
            if fill:
                for cell in row:
                    cell.fill = fill

    def generate_scorecard_excel(self, indicators: list[dict[str, Any]], summary: dict[str, Any], org_unit: str, org_unit_name: Optional[str], period: str) -> bytes:
        oxl = self._load()
        workbook = oxl["Workbook"]()
        self._write_sheet(
            workbook,
            "Summary",
            ["Field", "Value"],
            [
                ["Report", "WHO Validation Scorecard"],
                ["Organisation Unit", org_unit_name or org_unit],
                ["Period", period],
                ["Indicators reviewed", summary.get("total", 0)],
                ["Meeting target", summary.get("meeting_target", 0)],
                ["Score", f"{summary.get('score_pct', 0):.0f}%"],
            ],
            oxl,
        )
        self._write_sheet(
            workbook,
            "Indicators",
            ["Indicator", "ID", "Value", "Numerator", "Denominator", "Target", "Status"],
            [
                [
                    indicator.get("name", ""),
                    indicator.get("id", ""),
                    indicator.get("formatted_value", "N/A"),
                    indicator.get("numerator_value"),
                    indicator.get("denominator_value"),
                    indicator.get("target"),
                    indicator.get("status", "unknown"),
                ]
                for indicator in indicators
            ],
            oxl,
        )
        # Apply traffic-light fills to the Indicators sheet (status is col 7)
        indicators_sheet = workbook["Indicators"]
        self._apply_status_fills(indicators_sheet, 7, oxl)

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def generate_cascade_excel(self, cascade_type: str, steps: list[dict[str, Any]], org_unit: str, org_unit_name: Optional[str], period: str) -> bytes:
        oxl = self._load()
        workbook = oxl["Workbook"]()
        self._write_sheet(
            workbook,
            f"{cascade_type.upper()} Cascade",
            ["Step", "Indicator", "Count", "Coverage"],
            [[index, step.get("name", ""), step.get("count"), step.get("formatted_value", "N/A")] for index, step in enumerate(steps, start=1)],
            oxl,
        )
        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def generate_supply_excel(self, commodities: list[dict[str, Any]], org_unit: str, org_unit_name: Optional[str], period: str) -> bytes:
        oxl = self._load()
        workbook = oxl["Workbook"]()
        self._write_sheet(
            workbook,
            "Supply Status",
            ["Commodity", "Stock on hand", "Consumed", "Stockout days", "Days of use", "Status"],
            [
                [
                    commodity.get("commodity", ""),
                    commodity.get("stock_on_hand"),
                    commodity.get("consumed"),
                    commodity.get("stockout_days"),
                    commodity.get("days_of_use"),
                    commodity.get("status", "unknown"),
                ]
                for commodity in commodities
            ],
            oxl,
        )
        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()


class CSVExporter:
    """Generate CSV reports in memory."""

    def _to_bytes(self, rows: list[list[Any]]) -> bytes:
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerows(rows)
        return buffer.getvalue().encode("utf-8")

    def generate_scorecard_csv(self, indicators: list[dict[str, Any]], summary: dict[str, Any], org_unit: str, org_unit_name: Optional[str], period: str) -> bytes:
        rows = [
            ["Report", "WHO Validation Scorecard"],
            ["Organisation Unit", org_unit_name or org_unit],
            ["Period", period],
            ["Indicators reviewed", summary.get("total", 0)],
            ["Meeting target", summary.get("meeting_target", 0)],
            ["Score", f"{summary.get('score_pct', 0):.0f}%"],
            [],
            ["Indicator", "ID", "Value", "Numerator", "Denominator", "Target", "Status"],
        ]
        rows.extend(
            [
                indicator.get("name", ""),
                indicator.get("id", ""),
                indicator.get("formatted_value", "N/A"),
                indicator.get("numerator_value"),
                indicator.get("denominator_value"),
                indicator.get("target"),
                indicator.get("status", "unknown"),
            ]
            for indicator in indicators
        )
        return self._to_bytes(rows)

    def generate_cascade_csv(self, cascade_type: str, steps: list[dict[str, Any]], org_unit: str, org_unit_name: Optional[str], period: str) -> bytes:
        rows = [["Report", f"{cascade_type.upper()} Cascade"], ["Organisation Unit", org_unit_name or org_unit], ["Period", period], [], ["Step", "Indicator", "Count", "Coverage"]]
        rows.extend([index, step.get("name", ""), step.get("count"), step.get("formatted_value", "N/A")] for index, step in enumerate(steps, start=1))
        return self._to_bytes(rows)

    def generate_supply_csv(self, commodities: list[dict[str, Any]], org_unit: str, org_unit_name: Optional[str], period: str) -> bytes:
        rows = [["Report", "Supply Chain Status"], ["Organisation Unit", org_unit_name or org_unit], ["Period", period], [], ["Commodity", "Stock on hand", "Consumed", "Stockout days", "Days of use", "Status"]]
        rows.extend(
            [
                commodity.get("commodity", ""),
                commodity.get("stock_on_hand"),
                commodity.get("consumed"),
                commodity.get("stockout_days"),
                commodity.get("days_of_use"),
                commodity.get("status", "unknown"),
            ]
            for commodity in commodities
        )
        return self._to_bytes(rows)


class ExportService:
    """Facade for export generation and filename/content-type helpers."""

    def __init__(self) -> None:
        self.pdf_exporter = PDFExporter()
        self.excel_exporter = ExcelExporter()
        self.csv_exporter = CSVExporter()

    def get_filename(self, report_type: str, org_unit: str, period: str, format_name: str) -> str:
        return f"pmtct_{report_type}_{_safe_name(org_unit)}_{period}_{_timestamp()}.{format_name}"

    def get_content_type(self, format_name: str) -> str:
        return {
            "pdf": "application/pdf",
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "csv": "text/csv; charset=utf-8",
        }[format_name]

    def export_scorecard(self, format_name: str, indicators: list[dict[str, Any]], summary: dict[str, Any], org_unit: str, org_unit_name: Optional[str], period: str) -> bytes:
        if format_name == "pdf":
            return self.pdf_exporter.generate_scorecard_pdf(indicators, summary, org_unit, org_unit_name, period)
        if format_name == "xlsx":
            return self.excel_exporter.generate_scorecard_excel(indicators, summary, org_unit, org_unit_name, period)
        if format_name == "csv":
            return self.csv_exporter.generate_scorecard_csv(indicators, summary, org_unit, org_unit_name, period)
        raise ValueError(f"Unsupported export format: {format_name}")

    def export_cascade(self, format_name: str, cascade_type: str, steps: list[dict[str, Any]], org_unit: str, org_unit_name: Optional[str], period: str) -> bytes:
        if format_name == "pdf":
            return self.pdf_exporter.generate_cascade_pdf(cascade_type, steps, org_unit, org_unit_name, period)
        if format_name == "xlsx":
            return self.excel_exporter.generate_cascade_excel(cascade_type, steps, org_unit, org_unit_name, period)
        if format_name == "csv":
            return self.csv_exporter.generate_cascade_csv(cascade_type, steps, org_unit, org_unit_name, period)
        raise ValueError(f"Unsupported export format: {format_name}")

    def export_supply(self, format_name: str, commodities: list[dict[str, Any]], org_unit: str, org_unit_name: Optional[str], period: str) -> bytes:
        if format_name == "pdf":
            return self.pdf_exporter.generate_supply_pdf(commodities, org_unit, org_unit_name, period)
        if format_name == "xlsx":
            return self.excel_exporter.generate_supply_excel(commodities, org_unit, org_unit_name, period)
        if format_name == "csv":
            return self.csv_exporter.generate_supply_csv(commodities, org_unit, org_unit_name, period)
        raise ValueError(f"Unsupported export format: {format_name}")
